import sys
import firebase_admin
from firebase_admin import credentials, firestore
import openpyxl

def main():
    if len(sys.argv) < 2:
        print("Usage: export_script.py <taskId>")
        sys.exit(1)
    task_id = sys.argv[1]

    # Initialize Firebase Admin SDK (ensure you have serviceAccountKey.json in assets)
    cred = credentials.Certificate('assets/serviceAccountKey.json')
    firebase_admin.initialize_app(cred)
    db = firestore.client()

    # Load the template
    wb = openpyxl.load_workbook('assets/templateexport.xlsx')
    ws = wb.active

    # Fetch submissions for the task
    submissions = db.collection('submissions').where('taskId', '==', task_id).stream()
    row = 12  # Excel rows are 1-based
    for sub in submissions:
        data = sub.to_dict()
        # Example: write schooluid and submittedAt, add more fields as needed
        ws[f'A{row}'] = data.get('schooluid', '')
        ws[f'B{row}'] = str(data.get('submittedAt', ''))
        # Add more columns as needed...
        row += 1

    # Save to a new file
    out_path = f'assets/export_{task_id}.xlsx'
    wb.save(out_path)
    print(out_path)

if __name__ == '__main__':
    main()
